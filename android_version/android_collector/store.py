from __future__ import annotations

import csv
import json
import math
import re
from collections.abc import Callable
from dataclasses import replace
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook

from .driver import AndroidDriver
from .models import EvidencePaths, Metric, ObservedProfile, ObservedReel
from .ui_parser import parse_visible_reel
from .xlsx_style import apply_python_compatible_xlsx_style


# Keep public Android files interoperable with python_version. Android-only
# source details remain in .collector/android_observations.json.
CSV_FIELDS = (
    "collected_at",
    "url",
    "user_id",
    "username",
    "title",
    "hashtags",
    "audio_name",
    "location_name",
    "ad",
    "uploaded_at",
    "video_duration_seconds",
    "days_since_upload",
    "view_count",
    "like_count",
    "comment_count",
    "repost_count",
    "follower_count",
)
REEL_CHANGE_METRICS = (
    "view_count",
    "like_count",
    "comment_count",
    "repost_count",
    "follower_count",
)
REEL_CHANGE_FIELDS = (
    *(f"{field}_change" for field in REEL_CHANGE_METRICS),
    "reaction_rate_change",
)
ROW_COLLECTION_FIELDS = (
    "collection_number",
    "days_since_previous",
    *CSV_FIELDS,
    "reaction_rate",
    *REEL_CHANGE_FIELDS,
)
USER_COLLECTION_FIELDS = (
    "collection_number",
    "days_since_previous",
    "user_id",
    "username",
    "biography",
    "profile_category",
    "post_count",
    "following_count",
    "follower_count",
    "follower_count_change",
    "collected_at",
    "account_country",
)

METRIC_COLUMNS = (
    "view_count",
    "like_count",
    "comment_count",
    "share_count",
    "repost_count",
    "save_count",
    "likes_and_plays_count",
)
INTERNAL_DIRECTORY = ".collector"
INTERNAL_OBSERVATIONS_FILENAME = "android_observations.json"
HASHTAG_PATTERN = re.compile(r"#[\w]+", re.UNICODE)
INTEGER_PATTERN = re.compile(r"\d+")
SPONSORED_HASHTAG = "#협찬"
KST = timezone(timedelta(hours=9))


def is_instagram_reel_url(value: object) -> bool:
    candidate = str(value or "").strip()
    parsed = urlparse(candidate)
    return (
        parsed.scheme in {"http", "https"}
        and parsed.hostname is not None
        and parsed.hostname.casefold().endswith("instagram.com")
        and parsed.path.casefold().startswith(("/reel/", "/reels/"))
    )


def reel_url_identity(value: object) -> str:
    """Return a stable Reel identity while retaining the original shared URL."""
    candidate = str(value or "").strip()
    if not is_instagram_reel_url(candidate):
        return ""
    parsed = urlparse(candidate)
    path = parsed.path.rstrip("/") + "/"
    return f"{parsed.scheme.casefold()}://{parsed.hostname.casefold()}{path}"


def read_reel_urls_from_xlsx(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        headers = [str(value or "").strip().casefold() for value in next(rows, ())]
        url_index = next((index for index, value in enumerate(headers) if value in {"url", "reel_url"}), 0)
        urls: list[str] = []
        for row in rows:
            value = row[url_index] if len(row) > url_index else ""
            candidate = str(value or "").strip()
            if is_instagram_reel_url(candidate):
                urls.append(candidate)
        return urls
    finally:
        workbook.close()


def _replace_atomically(destination: Path, writer: Callable[[Path], None]) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f"{destination.stem}.tmp{destination.suffix}")
    try:
        writer(temporary)
        temporary.replace(destination)
    finally:
        if temporary.exists():
            temporary.unlink()


def _parse_datetime(value: object) -> datetime | None:
    try:
        parsed = datetime.fromisoformat(str(value or "").strip().replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed.astimezone(timezone.utc)


def _exact_count(value: object) -> int | str:
    if isinstance(value, int) and not isinstance(value, bool) and value >= 0:
        return value
    text = str(value or "").strip().replace(",", "")
    return int(text) if INTEGER_PATTERN.fullmatch(text) else ""


def _caption_without_hashtags(value: object) -> str:
    return re.sub(r"\s+", " ", HASHTAG_PATTERN.sub(" ", str(value or ""))).strip()


def _truncate_caption(value: str, maximum: int = 300) -> str:
    return value if len(value) <= maximum else f"{value[:maximum]}..."


def _extract_hashtags(value: object) -> str:
    unique: dict[str, str] = {}
    for hashtag in HASHTAG_PATTERN.findall(str(value or "")):
        unique.setdefault(hashtag.casefold(), hashtag)
    return " ".join(unique.values())


def _ad_status(raw: dict[str, object], caption: str) -> str:
    """Keep explicit in-app ads distinct from creator-declared sponsorship."""
    if raw.get("is_ad") is True:
        return "true"
    if any(tag.casefold() == SPONSORED_HASHTAG for tag in HASHTAG_PATTERN.findall(caption)):
        return "협찬"
    return "false"


def _reaction_rate(row: dict[str, object]) -> float | str:
    views = _exact_count(row.get("view_count"))
    followers = _exact_count(row.get("follower_count"))
    if not isinstance(views, int) or not isinstance(followers, int) or followers <= 0:
        return ""
    return views / followers


def _derived_fields(current: dict[str, object], previous: dict[str, object] | None) -> dict[str, object]:
    rate = _reaction_rate(current)
    derived: dict[str, object] = {"reaction_rate": rate}
    if previous is None:
        return {**derived, **{field: "" for field in REEL_CHANGE_FIELDS}}
    for field in REEL_CHANGE_METRICS:
        now = _exact_count(current.get(field))
        before = _exact_count(previous.get(field))
        derived[f"{field}_change"] = now - before if isinstance(now, int) and isinstance(before, int) else ""
    previous_rate = _reaction_rate(previous)
    derived["reaction_rate_change"] = (
        rate - previous_rate
        if isinstance(rate, float) and isinstance(previous_rate, float)
        else ""
    )
    return derived


def _days_since(previous: object, current: object) -> float | str:
    previous_time = _parse_datetime(previous)
    current_time = _parse_datetime(current)
    if previous_time is None or current_time is None:
        return ""
    elapsed = max(0.0, (current_time - previous_time).total_seconds() / 86_400)
    return math.floor(elapsed * 100 + 0.5) / 100


def _days_since_upload(collected_at: object, uploaded_at: object) -> int | str:
    """Return Korean-calendar days between the displayed upload date and capture."""
    collected_time = _parse_datetime(collected_at)
    try:
        uploaded_date = datetime.fromisoformat(str(uploaded_at or "").strip()).date()
    except ValueError:
        return ""
    if collected_time is None:
        return ""
    return max(0, (collected_time.astimezone(KST).date() - uploaded_date).days)


def _json_field_value(field: str, value: object) -> object | None:
    if value in (None, ""):
        return None
    if field == "ad":
        normalized = str(value).strip().casefold()
        return normalized == "true" if normalized in {"true", "false"} else value
    integer_fields = {
        "collection_number",
        "days_since_upload",
        "post_count",
        *REEL_CHANGE_METRICS,
        *(f"{item}_change" for item in REEL_CHANGE_METRICS),
    }
    if field in integer_fields:
        parsed = _exact_count(value)
        return parsed if isinstance(parsed, int) else value
    if field in {"days_since_previous", "reaction_rate", "reaction_rate_change"}:
        try:
            return float(str(value))
        except (TypeError, ValueError):
            return value
    return value


class CollectionStore:
    """Own Android evidence while exporting the python_version Reel history schema."""

    def __init__(
        self,
        data_dir: Path,
        *,
        reel_stem: str = "reels",
        user_stem: str = "users",
    ) -> None:
        self.data_dir = data_dir.resolve()
        self.reel_stem = reel_stem
        self.user_stem = user_stem
        self.evidence_dir = self.data_dir / "evidence"
        internal_name = (
            INTERNAL_OBSERVATIONS_FILENAME
            if reel_stem == "reels"
            else f"{reel_stem}_android_observations.json"
        )
        self.internal_path = self.data_dir / INTERNAL_DIRECTORY / internal_name
        self.rows = self._load_existing_rows()

    @staticmethod
    def _load_json_rows(path: Path) -> list[dict[str, object]]:
        if not path.exists():
            return []
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return []
        return [dict(row) for row in loaded if isinstance(row, dict)] if isinstance(loaded, list) else []

    def _load_existing_rows(self) -> list[dict[str, object]]:
        internal = self._load_json_rows(self.internal_path)
        if internal:
            return internal
        legacy_rows = self._load_json_rows(self.data_dir / f"{self.reel_stem}.json")
        if any("source_mode" in row or "reel_fingerprint" in row for row in legacy_rows):
            return legacy_rows
        return [self._public_row_to_raw(row, index) for index, row in enumerate(legacy_rows, start=1)]

    @staticmethod
    def _public_row_to_raw(row: dict[str, object], index: int) -> dict[str, object]:
        public_record = {field: row.get(field, "") for field in CSV_FIELDS}
        return {
            "source_mode": "imported",
            "source_query": "",
            "reel_url": public_record["url"],
            "reel_fingerprint": f"imported-{index}",
            "collected_at": public_record["collected_at"],
            "username": public_record["username"],
            "caption": public_record["title"],
            "audio_name": public_record["audio_name"],
            "public_record": public_record,
            "status": "imported",
        }

    def next_index(self) -> int:
        return len(self.rows) + 1

    def known_reel_fingerprints(self) -> set[str]:
        return {
            str(row.get("reel_fingerprint", "") or "").strip()
            for row in self.rows
            if str(row.get("reel_fingerprint", "") or "").strip()
        }

    def latest_raw_for_reel_url(self, reel_url: str) -> dict[str, object] | None:
        """Return the latest saved observation for a canonical shared Reel URL."""
        identity = reel_url_identity(reel_url)
        if not identity:
            return None
        matches = [
            row
            for row in self.rows
            if reel_url_identity(row.get("reel_url", "")) == identity
        ]
        if not matches:
            return None
        return max(
            matches,
            key=lambda row: _parse_datetime(row.get("collected_at")) or datetime.min.replace(tzinfo=timezone.utc),
        )

    def preserve_refresh_fields(self, observation: ObservedReel) -> ObservedReel:
        """Keep prior static/public values when a refresh screen omits them.

        A re-opened Reel can temporarily omit a caption, a location, or a
        profile field while still exposing its current metrics.  The refresh
        record therefore preserves the latest known value for blank static
        fields and fills missing metrics from the prior observation; newly
        rendered values always win.
        """
        previous = self.latest_raw_for_reel_url(observation.reel_url)
        if previous is None:
            return observation

        previous_metrics: dict[str, Metric] = {}
        for key in METRIC_COLUMNS:
            value = _exact_count(previous.get(key, ""))
            if isinstance(value, int):
                previous_metrics[key] = Metric(
                    label=key.removesuffix("_count"),
                    value=value,
                    raw_text=str(previous.get(f"{key}_raw", "") or value),
                )
        previous_post_count = _exact_count(previous.get("post_count", ""))
        previous_following_count = _exact_count(previous.get("following_count", ""))
        previous_follower_count = _exact_count(previous.get("follower_count", ""))
        old_profile = ObservedProfile(
            user_id=str(previous.get("user_id", "") or ""),
            username=str(previous.get("username", "") or ""),
            biography=str(previous.get("biography", "") or ""),
            profile_category=str(previous.get("profile_category", "") or ""),
            account_country=str(previous.get("account_country", "") or ""),
            post_count=previous_post_count if isinstance(previous_post_count, int) else None,
            following_count=previous_following_count if isinstance(previous_following_count, int) else None,
            follower_count=previous_follower_count if isinstance(previous_follower_count, int) else None,
        )
        profile = observation.profile
        merged_profile = replace(
            profile,
            user_id=profile.user_id or old_profile.user_id,
            username=profile.username or old_profile.username,
            biography=profile.biography or old_profile.biography,
            profile_category=profile.profile_category or old_profile.profile_category,
            account_country=profile.account_country or old_profile.account_country,
            post_count=profile.post_count if profile.post_count is not None else old_profile.post_count,
            following_count=profile.following_count if profile.following_count is not None else old_profile.following_count,
            follower_count=profile.follower_count if profile.follower_count is not None else old_profile.follower_count,
        )
        previous_private = previous.get("like_count_is_private")
        return replace(
            observation,
            username=observation.username or str(previous.get("username", "") or ""),
            caption=observation.caption or str(previous.get("caption", "") or ""),
            audio_name=observation.audio_name or str(previous.get("audio_name", "") or ""),
            location_name=observation.location_name or str(previous.get("location_name", "") or ""),
            uploaded_at=observation.uploaded_at or str(previous.get("uploaded_at", "") or ""),
            is_ad=observation.is_ad or previous.get("is_ad") is True,
            profile=merged_profile,
            metrics={**previous_metrics, **observation.metrics},
            like_count_is_private=(
                observation.like_count_is_private
                if observation.like_count_is_private is not None
                else previous_private if isinstance(previous_private, bool) else None
            ),
        )

    def save_evidence(
        self,
        index: int,
        xml: str,
        driver: AndroidDriver,
        *,
        suffix: str = "",
        capture_screenshot: bool = True,
    ) -> EvidencePaths:
        self.evidence_dir.mkdir(parents=True, exist_ok=True)
        xml_path = self.evidence_dir / f"{index:06d}{suffix}.xml"
        png_path = self.evidence_dir / f"{index:06d}{suffix}.png"
        xml_path.write_text(xml, encoding="utf-8")
        if capture_screenshot:
            driver.capture_screenshot(png_path)
            return EvidencePaths(xml_path=str(xml_path), png_path=str(png_path))
        return EvidencePaths(xml_path=str(xml_path), png_path="")

    @staticmethod
    def _row(observation: ObservedReel) -> dict[str, object]:
        row: dict[str, object] = {
            "collected_at": observation.collected_at,
            "source_mode": observation.source_mode,
            "source_query": observation.source_query,
            "reel_url": observation.reel_url,
            "reel_fingerprint": observation.reel_fingerprint,
            "username": observation.username,
            "caption": observation.caption,
            "audio_name": observation.audio_name,
            "location_name": observation.location_name,
            "uploaded_at": observation.uploaded_at,
            "is_ad": observation.is_ad,
            "user_id": observation.profile.user_id,
            "biography": observation.profile.biography,
            "profile_category": observation.profile.profile_category,
            "account_country": observation.profile.account_country,
            "post_count": observation.profile.post_count if observation.profile.post_count is not None else "",
            "following_count": observation.profile.following_count if observation.profile.following_count is not None else "",
            "follower_count": observation.profile.follower_count if observation.profile.follower_count is not None else "",
            "visible_metrics_json": json.dumps(observation.visible_metrics, ensure_ascii=False, sort_keys=True),
            "evidence_xml_path": observation.evidence_paths.xml_path,
            "evidence_png_path": observation.evidence_paths.png_path,
            "detail_evidence_xml_path": observation.detail_evidence_paths.xml_path,
            "detail_evidence_png_path": observation.detail_evidence_paths.png_path,
            "like_count_is_private": observation.like_count_is_private,
            "status": observation.status,
        }
        for column in METRIC_COLUMNS:
            metric = observation.metrics.get(column)
            row[column] = metric.value if metric is not None else ""
            row[f"{column}_raw"] = metric.raw_text if metric is not None else ""
        return row

    @staticmethod
    def _public_base_row(raw: dict[str, object]) -> dict[str, object]:
        imported = raw.get("public_record")
        if isinstance(imported, dict):
            return {field: imported.get(field, "") for field in CSV_FIELDS}
        caption = str(raw.get("caption", "") or "")
        return {
            "collected_at": raw.get("collected_at", ""),
            "url": raw.get("reel_url", ""),
            "user_id": "",
            "username": raw.get("username", ""),
            "title": _truncate_caption(_caption_without_hashtags(caption)),
            "hashtags": _extract_hashtags(caption),
            "audio_name": raw.get("audio_name", ""),
            "location_name": raw.get("location_name", ""),
            "ad": _ad_status(raw, caption),
            "uploaded_at": raw.get("uploaded_at", ""),
            "video_duration_seconds": "",
            "days_since_upload": _days_since_upload(raw.get("collected_at"), raw.get("uploaded_at")),
            "view_count": _exact_count(raw.get("view_count")),
            "like_count": _exact_count(raw.get("like_count")),
            "comment_count": _exact_count(raw.get("comment_count")),
            "repost_count": _exact_count(raw.get("repost_count")),
            "follower_count": "",
        }

    @staticmethod
    def _identity(raw: dict[str, object], index: int) -> str:
        url = reel_url_identity(raw.get("reel_url", ""))
        if url:
            return f"url:{url}"
        fingerprint = str(raw.get("reel_fingerprint", "") or "").strip()
        return f"fingerprint:{fingerprint}" if fingerprint else f"row:{index}"

    def _public_rows(self) -> list[dict[str, object]]:
        output: list[dict[str, object]] = [{} for _ in self.rows]
        ordered_indexes = sorted(
            range(len(self.rows)),
            key=lambda index: (
                self._identity(self.rows[index], index),
                _parse_datetime(self.rows[index].get("collected_at")) or datetime.min.replace(tzinfo=timezone.utc),
                index,
            ),
        )
        previous_by_identity: dict[str, dict[str, object]] = {}
        collection_count: dict[str, int] = {}
        for index in ordered_indexes:
            raw = self.rows[index]
            identity = self._identity(raw, index)
            current = self._public_base_row(raw)
            previous = previous_by_identity.get(identity)
            collection_count[identity] = collection_count.get(identity, 0) + 1
            public = {
                "collection_number": collection_count[identity],
                "days_since_previous": _days_since(
                    previous.get("collected_at", "") if previous else "",
                    current.get("collected_at", ""),
                ),
                **current,
                **_derived_fields(current, previous),
            }
            output[index] = {field: public.get(field, "") for field in ROW_COLLECTION_FIELDS}
            previous_by_identity[identity] = current
        return output

    @staticmethod
    def _user_base_row(raw: dict[str, object]) -> dict[str, object]:
        imported = raw.get("public_record")
        source = imported if isinstance(imported, dict) else raw
        return {
            "user_id": source.get("user_id", ""),
            "username": source.get("username", ""),
            "biography": source.get("biography", ""),
            "profile_category": source.get("profile_category", ""),
            "account_country": source.get("account_country", ""),
            "post_count": _exact_count(source.get("post_count")),
            "following_count": _exact_count(source.get("following_count")),
            "follower_count": _exact_count(source.get("follower_count")),
            "collected_at": source.get("collected_at", ""),
        }

    @staticmethod
    def _user_identity(row: dict[str, object], index: int) -> str:
        user_id = str(row.get("user_id", "") or "").strip()
        if user_id:
            return f"id:{user_id}"
        username = str(row.get("username", "") or "").strip().casefold()
        return f"username:{username}" if username else f"row:{index}"

    def _user_rows(self) -> list[dict[str, object]]:
        candidates = [
            (index, self._user_base_row(raw))
            for index, raw in enumerate(self.rows)
            if str(self._user_base_row(raw).get("user_id", "") or "").strip()
            or str(self._user_base_row(raw).get("username", "") or "").strip()
        ]
        output: list[dict[str, object]] = [{} for _ in candidates]
        ordered = sorted(
            enumerate(candidates),
            key=lambda item: (
                self._user_identity(item[1][1], item[1][0]),
                _parse_datetime(item[1][1].get("collected_at")) or datetime.min.replace(tzinfo=timezone.utc),
                item[1][0],
            ),
        )
        previous_by_identity: dict[str, dict[str, object]] = {}
        collection_count: dict[str, int] = {}
        for output_index, (source_index, current) in ordered:
            identity = self._user_identity(current, source_index)
            previous = previous_by_identity.get(identity)
            collection_count[identity] = collection_count.get(identity, 0) + 1
            follower_count = _exact_count(current.get("follower_count"))
            previous_follower_count = _exact_count(previous.get("follower_count")) if previous else ""
            user = {
                "collection_number": collection_count[identity],
                "days_since_previous": _days_since(
                    previous.get("collected_at", "") if previous else "",
                    current.get("collected_at", ""),
                ),
                **current,
                "follower_count_change": (
                    follower_count - previous_follower_count
                    if isinstance(follower_count, int) and isinstance(previous_follower_count, int)
                    else ""
                ),
            }
            output[output_index] = {field: user.get(field, "") for field in USER_COLLECTION_FIELDS}
            previous_by_identity[identity] = current
        return output

    def append(self, observation: ObservedReel) -> None:
        self.rows.append(self._row(observation))

    def _backfill_audio_names_from_evidence(self) -> None:
        """Fill legacy empty audio fields from their retained visible XML."""
        for raw in self.rows:
            if str(raw.get("audio_name", "") or "").strip():
                continue
            evidence_value = str(raw.get("evidence_xml_path", "") or "").strip()
            if not evidence_value:
                continue
            evidence_path = Path(evidence_value)
            if not evidence_path.is_absolute():
                evidence_path = self.data_dir / evidence_path
            try:
                evidence_xml = evidence_path.read_text(encoding="utf-8")
            except OSError:
                continue
            observed = parse_visible_reel(
                evidence_xml,
                source_mode=str(raw.get("source_mode", "") or ""),
                source_query=str(raw.get("source_query", "") or ""),
                reel_url=str(raw.get("reel_url", "") or ""),
                collected_at=str(raw.get("collected_at", "") or ""),
            )
            if observed.audio_name:
                raw["audio_name"] = observed.audio_name

    def export(self) -> None:
        self._backfill_audio_names_from_evidence()
        rows = self._public_rows()
        user_rows = self._user_rows()

        def write_csv(path: Path) -> None:
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=ROW_COLLECTION_FIELDS, quoting=csv.QUOTE_ALL)
                writer.writeheader()
                writer.writerows(rows)

        def write_json(path: Path) -> None:
            payload = [
                {field: _json_field_value(field, row.get(field, "")) for field in ROW_COLLECTION_FIELDS}
                for row in rows
            ]
            path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        def write_users_csv(path: Path) -> None:
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=USER_COLLECTION_FIELDS, quoting=csv.QUOTE_ALL)
                writer.writeheader()
                writer.writerows(user_rows)

        def write_users_json(path: Path) -> None:
            payload = [
                {field: _json_field_value(field, row.get(field, "")) for field in USER_COLLECTION_FIELDS}
                for row in user_rows
            ]
            path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        def write_xlsx(path: Path) -> None:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "reels"
            worksheet.append(list(ROW_COLLECTION_FIELDS))
            for row in rows:
                worksheet.append([row[field] for field in ROW_COLLECTION_FIELDS])
            apply_python_compatible_xlsx_style(workbook)
            workbook.save(path)
            workbook.close()

        def write_users_xlsx(path: Path) -> None:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "users"
            worksheet.append(list(USER_COLLECTION_FIELDS))
            for row in user_rows:
                worksheet.append([row[field] for field in USER_COLLECTION_FIELDS])
            apply_python_compatible_xlsx_style(workbook)
            workbook.save(path)
            workbook.close()

        def write_internal(path: Path) -> None:
            path.write_text(json.dumps(self.rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        _replace_atomically(self.data_dir / f"{self.reel_stem}.csv", write_csv)
        _replace_atomically(self.data_dir / f"{self.reel_stem}.json", write_json)
        _replace_atomically(self.data_dir / f"{self.reel_stem}.xlsx", write_xlsx)
        _replace_atomically(self.data_dir / f"{self.user_stem}.csv", write_users_csv)
        _replace_atomically(self.data_dir / f"{self.user_stem}.json", write_users_json)
        _replace_atomically(self.data_dir / f"{self.user_stem}.xlsx", write_users_xlsx)
        _replace_atomically(self.internal_path, write_internal)
