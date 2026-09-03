"""Python-version-compatible presentation rules for Android export workbooks."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Aptos", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATE_ALIGNMENT = Alignment(vertical="center")
KST = timezone(timedelta(hours=9))

DATE_FIELDS = {"collected_at", "uploaded_at"}
NUMERIC_FIELDS = {
    "collection_number",
    "days_since_upload",
    "video_duration_seconds",
    "view_count",
    "view_count_change",
    "post_count",
    "following_count",
    "follower_count",
    "follower_count_change",
    "like_count",
    "like_count_change",
    "comment_count",
    "comment_count_change",
    "repost_count",
    "repost_count_change",
}
PERCENT_FIELDS = {"reaction_rate", "reaction_rate_change"}
SIGNED_FIELDS = {
    "view_count_change",
    "like_count_change",
    "comment_count_change",
    "repost_count_change",
    "follower_count_change",
}
TEXT_IDENTIFIER_FIELDS = {"user_id"}


def _base_field(field_name: object) -> str:
    return str(field_name or "").strip()


def _parse_utc_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    try:
        parsed = datetime.fromisoformat(str(value or "").strip().replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed


def _display_length(value: object) -> int:
    if isinstance(value, datetime):
        return 19
    return len(str(value or ""))


def _column_width(values: Iterable[object], header: str) -> float:
    longest = max([len(header), *(_display_length(value) for value in values)], default=8)
    return min(max(longest + 2, 10), 42)


def apply_python_compatible_xlsx_style(workbook: Workbook) -> None:
    """Apply the same practical workbook presentation as python_version output."""
    for worksheet in workbook.worksheets:
        if worksheet.max_row < 1 or worksheet.max_column < 1:
            continue
        headers = [str(worksheet.cell(1, column).value or "") for column in range(1, worksheet.max_column + 1)]
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        worksheet.row_dimensions[1].height = 30

        for column, header in enumerate(headers, start=1):
            base_field = _base_field(header)
            header_cell = worksheet.cell(1, column)
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.alignment = HEADER_ALIGNMENT
            values: list[object] = []
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, column)
                if base_field in DATE_FIELDS:
                    parsed = _parse_utc_datetime(cell.value)
                    if parsed is not None:
                        if base_field == "uploaded_at":
                            cell.value = parsed.date()
                            cell.number_format = "yyyy-mm-dd"
                        else:
                            cell.value = parsed.astimezone(KST).replace(tzinfo=None)
                            cell.number_format = "yyyy-mm-dd hh:mm:ss"
                        cell.alignment = DATE_ALIGNMENT
                elif base_field in PERCENT_FIELDS:
                    cell.number_format = "0.00%"
                elif base_field in SIGNED_FIELDS:
                    cell.number_format = "+#,##0;-#,##0;0"
                elif base_field in NUMERIC_FIELDS:
                    cell.number_format = "#,##0"
                elif base_field in TEXT_IDENTIFIER_FIELDS:
                    cell.number_format = "@"
                cell.font = BODY_FONT
                values.append(cell.value)
            worksheet.column_dimensions[get_column_letter(column)].width = _column_width(values[:200], header)
